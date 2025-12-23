import json
import os
import tempfile
import openpyxl
import win32com.client as win32
import win32api
from openpyxl.chart import Reference


def excel_processing():
    # Загрузка книги
    file_path = "project/assets/report.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb["Report"]

    for row in range(2, 62):
        ws[f'M{row}'] = None

    # Загрузка статистических данных из JSON
    with open("project/configuration/graph.json", "r", encoding="utf-8") as file:
        data = json.load(file)

    mean_points = data["mean_points"]
    reach_points = data["reach_points"]
    ws['F6'] = data["total_schemes"]
    ws['F8'] = data["checked_schemes"]
    ws['F10'] = data["good_schemes"]
    ws['F12'] = data["bad_schemes"]

    # Заполнение книги данными
    for row, value in enumerate(mean_points, start=2):
        ws[f'M{row}'] = float(value)

    for row, value in enumerate(reach_points, start=2):
        ws[f'O{row}'] = float(value)

    # Обновление графика
    for i, chart in enumerate(ws._charts):
        if isinstance(chart, openpyxl.chart.LineChart):
            # Удаление старых данных
            while len(chart.series) > 0:
                chart.series.pop()

            # Выбор столбца в зависимости от индекса графика
            if i == 0:
                data_col = 13
            else:
                data_col = 15

            # Добавление новых данных
            data = Reference(ws, min_col=data_col, min_row=1, max_row=61)
            chart.add_data(data, titles_from_data=True)
            categories = Reference(ws, min_col=1, min_row=2, max_row=61)
            chart.set_categories(categories)

            for series in chart.series:
                series.smooth = False

    # Сохранение изменений
    wb.save(file_path)

    # Выделение и печать диапазона A1:H47
    excel = win32.Dispatch('Excel.Application')
    excel.Visible = True  # Для отладки!
    wb = excel.Workbooks.Open(os.path.abspath(r"project/assets/report.xlsx"))
    ws = wb.Worksheets(1)

    # Создание временной книги и вставка туда содержимого основной книги
    temp_wb = excel.Workbooks.Add()
    temp_ws = temp_wb.Worksheets(1)
    ws.UsedRange.Copy()
    temp_ws.Range("A1").PasteSpecial(Paste=-4104)
    charts = ws.ChartObjects()

    # Диаграмма средних значений
    if charts.Count >= 1:
        charts(1).Copy()
        temp_ws.Paste()
        new_chart = temp_ws.ChartObjects(temp_ws.ChartObjects().Count)
        new_chart.Top = temp_ws.Range("A14").Top
        new_chart.Left = temp_ws.Range("A14").Left

    # Диаграмма размахов
    if charts.Count >= 2:
        charts(2).Copy()
        temp_ws.Paste()
        new_chart = temp_ws.ChartObjects(temp_ws.ChartObjects().Count)
        new_chart.Top = temp_ws.Range("A29").Top
        new_chart.Left = temp_ws.Range("A29").Left

    # Удаляем всего, что лежит за пределами нужного диапазона
    last_row = temp_ws.UsedRange.Rows.Count
    last_col = temp_ws.UsedRange.Columns.Count
    if last_row > 47:
        temp_ws.Range(f"A48:A{last_row}").EntireRow.Delete()
    if last_col > 7:
        temp_ws.Range(f"H1:{chr(64 + last_col)}1").EntireColumn.Delete()

    # Сохранение временного файла (с удалением предыдущей версии), закрытие Excel
    temp_dir = tempfile.gettempdir()
    temp_file = os.path.join(temp_dir, "temp_print.xlsx")

    if temp_file and os.path.exists(temp_file):
        os.remove(temp_file)

    temp_wb.SaveAs(temp_file, FileFormat=51)
    temp_wb.Close(SaveChanges=False)
    wb.Close(SaveChanges=False)
    excel.Quit()

    # Печать через ShellExecute
    # result = win32api.ShellExecute(
    #     0,        # hwnd
    #     "printto",  # operation
    #     temp_file,      # file
    #     None,    # parameters
    #     temp_dir,       # directory
    #     0        # show cmd
    # )
    #
    # print(f"Отправлено на печать: {temp_file}", result)
